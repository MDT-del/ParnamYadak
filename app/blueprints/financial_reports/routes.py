# ---------------------------------------------
# فایل: routes.py (financial_reports)
# توضیح: گزارش‌های مالی و تحلیلی سیستم (فروش، مشتریان، فاکتورها، کوپن‌ها)
# ---------------------------------------------

from flask import render_template, Blueprint, request, jsonify, current_app, flash, redirect, url_for
from flask_login import login_required, current_user
from sqlalchemy import func, desc, and_, extract
from app import db
from app.models import Order, Person, InventoryProduct, MechanicProfile, BotOrder, BotOrderItem, InStoreOrder, FinancialTransaction
from app.decorators import permission_required
import json
import jdatetime
import datetime
from datetime import timedelta

financial_reports_bp = Blueprint('financial_reports',
                                 __name__,
                                 template_folder='templates')


@financial_reports_bp.route('/')
@login_required
@permission_required('view_financial_reports')
def index():
    """
    صفحه اصلی گزارش‌های مالی با آمار کلی و نمودارهای زنده
    """
    
    # محاسبه آمار کلی
    today = datetime.datetime.now().date()
    start_of_month = datetime.datetime(today.year, today.month, 1)
    start_of_year = datetime.datetime(today.year, 1, 1)
    
    # آمار فروش کل (حضوری + ربات)
    total_sales_telegram = db.session.query(func.sum(Order.total_price)).filter(
        Order.status.in_(['تایید شده', 'پرداخت شده', 'تکمیل شده'])).scalar() or 0
    total_sales_instore = db.session.query(func.sum(InStoreOrder.total_price)).filter(
        InStoreOrder.status.in_(['تحویل داده شده', 'تکمیل شده'])).scalar() or 0
    total_sales_bot = db.session.query(func.sum(BotOrder.total_amount)).filter(
        BotOrder.status.in_(['تایید شده', 'پرداخت شده', 'تکمیل شده'])).scalar() or 0
    total_sales = total_sales_telegram + total_sales_instore + total_sales_bot
    
    # فروش ماه جاری
    monthly_sales_telegram = db.session.query(func.sum(Order.total_price)).filter(
        Order.status.in_(['تایید شده', 'پرداخت شده', 'تکمیل شده']),
        Order.order_date >= start_of_month).scalar() or 0
    monthly_sales_instore = db.session.query(func.sum(InStoreOrder.total_price)).filter(
        InStoreOrder.status.in_(['تحویل داده شده', 'تکمیل شده']),
        InStoreOrder.created_at >= start_of_month).scalar() or 0
    monthly_sales_bot = db.session.query(func.sum(BotOrder.total_amount)).filter(
        BotOrder.status.in_(['تایید شده', 'پرداخت شده', 'تکمیل شده']),
        BotOrder.created_at >= start_of_month).scalar() or 0
    monthly_sales = monthly_sales_telegram + monthly_sales_instore + monthly_sales_bot
    
    # فروش امروز
    today_sales_telegram = db.session.query(func.sum(Order.total_price)).filter(
        Order.status.in_(['تایید شده', 'پرداخت شده', 'تکمیل شده']),
        func.date(Order.order_date) == today).scalar() or 0
    today_sales_instore = db.session.query(func.sum(InStoreOrder.total_price)).filter(
        InStoreOrder.status.in_(['تحویل داده شده', 'تکمیل شده']),
        func.date(InStoreOrder.created_at) == today).scalar() or 0
    today_sales_bot = db.session.query(func.sum(BotOrder.total_amount)).filter(
        BotOrder.status.in_(['تایید شده', 'پرداخت شده', 'تکمیل شده']),
        func.date(BotOrder.created_at) == today).scalar() or 0
    today_sales = today_sales_telegram + today_sales_instore + today_sales_bot
    
    # تعداد سفارشات
    total_orders_telegram = db.session.query(func.count(Order.id)).scalar() or 0
    total_orders_instore = db.session.query(func.count(InStoreOrder.id)).scalar() or 0
    total_orders_bot = db.session.query(func.count(BotOrder.id)).scalar() or 0
    total_orders = total_orders_telegram + total_orders_instore + total_orders_bot
    
    # تعداد مشتریان
    total_customers = db.session.query(func.count(Person.id)).filter(Person.person_type == 'customer').scalar() or 0
    
    # تعداد محصولات (از انبار)
    total_products = db.session.query(func.count(InventoryProduct.id)).scalar() or 0
    
    # آمار بیعانه‌ها
    total_deposits = db.session.query(func.sum(InStoreOrder.deposit_amount)).filter(
        InStoreOrder.deposit_amount > 0).scalar() or 0
    
    # باقی‌مانده‌های پرداخت نشده
    total_remaining = db.session.query(func.sum(InStoreOrder.total_price - InStoreOrder.deposit_amount)).filter(
        InStoreOrder.deposit_amount > 0).scalar() or 0
    
    # محصولات پرفروش (۱۰ محصول برتر بر اساس مجموع فروش در ۶ ماه گذشته)
    six_months_ago = datetime.datetime.now() - timedelta(days=180)
    product_sales = {}
    # سفارشات حضوری
    instore_orders = InStoreOrder.query.filter(InStoreOrder.created_at >= six_months_ago).all()
    for order in instore_orders:
        try:
            items = json.loads(order.products_info)
        except Exception:
            items = []
        for item in items:
            name = item.get('name')
            qty = int(item.get('qty', 0) or 0)
            total = float(item.get('price', 0) or 0) * qty
            if name:
                if name not in product_sales:
                    product_sales[name] = {'total_qty': 0, 'total_revenue': 0}
                product_sales[name]['total_qty'] += qty
                product_sales[name]['total_revenue'] += total
    # سفارشات ربات
    bot_items = BotOrderItem.query.join(BotOrder).filter(BotOrder.created_at >= six_months_ago).all()
    for item in bot_items:
        name = item.product_name
        qty = int(item.quantity or 0)
        total = float(item.total_price or 0)
        if name:
            if name not in product_sales:
                product_sales[name] = {'total_qty': 0, 'total_revenue': 0}
            product_sales[name]['total_qty'] += qty
            product_sales[name]['total_revenue'] += total
    # ۱۰ محصول پرفروش
    top_products = sorted(product_sales.items(), key=lambda x: x[1]['total_revenue'], reverse=True)[:10]
    top_products = [(name, data['total_qty'], data['total_revenue']) for name, data in top_products]
    # آمار فروش 12 ماه گذشته برای نمودار
    sales_data = []
    for i in range(12):
        month = today.month - i
        year = today.year
        if month <= 0:
            month += 12
            year -= 1
        monthly_sale_telegram = db.session.query(func.sum(Order.total_price)).filter(
            func.extract('year', Order.order_date) == year,
            func.extract('month', Order.order_date) == month).scalar() or 0
        monthly_sale_instore = db.session.query(func.sum(InStoreOrder.total_price)).filter(
            func.extract('year', InStoreOrder.created_at) == year,
            func.extract('month', InStoreOrder.created_at) == month,
            InStoreOrder.status.in_(['تحویل داده شده', 'تکمیل شده'])).scalar() or 0
        monthly_sale_bot = db.session.query(func.sum(BotOrder.total_amount)).filter(
            func.extract('year', BotOrder.created_at) == year,
            func.extract('month', BotOrder.created_at) == month,
            BotOrder.status.in_(['تایید شده', 'پرداخت شده', 'تکمیل شده'])).scalar() or 0
        monthly_total = (monthly_sale_telegram or 0) + (monthly_sale_instore or 0) + (monthly_sale_bot or 0)
        gregorian_date_obj = datetime.date(year, month, 1)
        shamsi_label = jdatetime.date.fromgregorian(date=gregorian_date_obj).strftime("%Y/%m")
        sales_data.insert(0, {'month': shamsi_label, 'total': monthly_total})
    return render_template('financial_reports.html',
                           title="گزارش‌های مالی",
                           total_sales=total_sales,
                           monthly_sales=monthly_sales,
                           today_sales=today_sales,
                           total_orders=total_orders,
                           total_customers=total_customers,
                           total_products=total_products,
                           total_deposits=total_deposits,
                           total_remaining=total_remaining,
                           top_products=top_products,
                           sales_data=sales_data)


@financial_reports_bp.route('/invoice-report')
@login_required
@permission_required('view_financial_reports')
def invoice_report():
    """
    گزارش فاکتورها و وضعیت پرداخت‌ها - شامل سفارشات ربات و حضوری
    """
    from app.models import InStoreOrder, Person
    
    # فاکتورهای ربات (اخیر)
    recent_bot_invoices = BotOrder.query.filter(
        BotOrder.status.in_(['تایید شده', 'پرداخت شده', 'تکمیل شده'])).order_by(desc(
            BotOrder.created_at)).limit(20).all()

    # اضافه کردن customer_id برای سفارشات ربات
    for bot_order in recent_bot_invoices:
        if bot_order.customer_phone:
            customer = Person.query.filter_by(phone_number=bot_order.customer_phone, person_type='customer').first()
            bot_order.person_id = customer.id if customer else 1
        else:
            bot_order.person_id = 1

    # فاکتورهای حضوری (اخیر)
    recent_instore_invoices = InStoreOrder.query.filter(
        InStoreOrder.status.in_(['تحویل داده شده', 'تکمیل شده'])).order_by(desc(
            InStoreOrder.created_at)).limit(20).all()

    # آمار فاکتورهای ربات
    total_bot_invoices = BotOrder.query.filter(
        BotOrder.status.in_(['تایید شده', 'پرداخت شده', 'تکمیل شده'])).count()
    total_bot_amount = db.session.query(func.sum(
        BotOrder.total_amount)).filter(
            BotOrder.status.in_(['تایید شده', 'پرداخت شده', 'تکمیل شده'])).scalar() or 0
    paid_bot_invoices = BotOrder.query.filter(
        BotOrder.status == 'پرداخت شده').count()
    unpaid_bot_invoices = BotOrder.query.filter(
        BotOrder.status == 'در انتظار پرداخت').count()
    pending_bot_invoices = BotOrder.query.filter(
        BotOrder.status == 'در انتظار تایید پرداخت').count()

    # آمار فاکتورهای حضوری
    total_instore_invoices = InStoreOrder.query.filter(
        InStoreOrder.status.in_(['تحویل داده شده', 'تکمیل شده'])).count()
    total_instore_amount = db.session.query(func.sum(
        InStoreOrder.total_price)).filter(
            InStoreOrder.status.in_(['تحویل داده شده', 'تکمیل شده'])).scalar() or 0
    paid_instore_invoices = InStoreOrder.query.filter(
        InStoreOrder.status == 'تکمیل شده').count()
    unpaid_instore_invoices = InStoreOrder.query.filter(
        InStoreOrder.status == 'جدید').count()
    pending_instore_invoices = InStoreOrder.query.filter(
        InStoreOrder.status == 'در حال پردازش').count()

    # آمار کلی
    total_invoices = total_bot_invoices + total_instore_invoices
    total_invoice_amount = total_bot_amount + total_instore_amount
    total_paid_invoices = paid_bot_invoices + paid_instore_invoices
    total_unpaid_invoices = unpaid_bot_invoices + unpaid_instore_invoices
    total_pending_invoices = pending_bot_invoices + pending_instore_invoices

    # فاکتورهای پرداخت نشده (برای نمایش)
    unpaid_bot = BotOrder.query.filter(
        BotOrder.status == 'در انتظار پرداخت').order_by(desc(
            BotOrder.created_at)).limit(10).all()
    unpaid_instore = InStoreOrder.query.filter(
        InStoreOrder.status == 'جدید').order_by(desc(
            InStoreOrder.created_at)).limit(10).all()

    return render_template('invoice_report.html',
                           title="گزارش فاکتورها",
                           # فاکتورهای اخیر
                           recent_bot_invoices=recent_bot_invoices,
                           recent_instore_invoices=recent_instore_invoices,
                           # آمار کلی
                           total_invoices=total_invoices,
                           total_invoice_amount=total_invoice_amount,
                           total_paid_invoices=total_paid_invoices,
                           total_unpaid_invoices=total_unpaid_invoices,
                           total_pending_invoices=total_pending_invoices,
                           # آمار تفکیکی
                           total_bot_invoices=total_bot_invoices,
                           total_bot_amount=total_bot_amount,
                           paid_bot_invoices=paid_bot_invoices,
                           unpaid_bot_invoices=unpaid_bot_invoices,
                           pending_bot_invoices=pending_bot_invoices,
                           total_instore_invoices=total_instore_invoices,
                           total_instore_amount=total_instore_amount,
                           paid_instore_invoices=paid_instore_invoices,
                           unpaid_instore_invoices=unpaid_instore_invoices,
                           pending_instore_invoices=pending_instore_invoices,
                           # فاکتورهای پرداخت نشده
                           unpaid_bot=unpaid_bot,
                           unpaid_instore=unpaid_instore,
                           now=datetime.datetime.now())


@financial_reports_bp.route('/api/sales-chart')
@login_required
@permission_required('view_financial_reports')
def api_sales_chart():
    """API برای نمودار فروش"""
    period = request.args.get('period', 'monthly')
    days = int(request.args.get('days', 30))

    end_date = datetime.datetime.now()
    start_date = end_date - timedelta(days=days)

    if period == 'daily':
        data = db.session.query(
            func.date(Order.order_date).label('date'),
            func.sum(Order.final_price).label('total'),
            func.count(Order.id).label('count')).filter(
                Order.order_date.between(start_date, end_date)).group_by(
                    func.date(Order.order_date)).order_by(
                        func.date(Order.order_date)).all()

        result = [{
            'date': item.date.strftime('%Y-%m-%d'),
            'total': float(item.total),
            'count': item.count
        } for item in data]
    else:
        data = db.session.query(
            func.extract('year', Order.order_date).label('year'),
            func.extract('month', Order.order_date).label('month'),
            func.sum(Order.final_price).label('total'),
            func.count(Order.id).label('count')).filter(
                Order.order_date.between(start_date, end_date)).group_by(
                    func.extract('year', Order.order_date),
                    func.extract('month', Order.order_date)).order_by(
                        func.extract('year', Order.order_date),
                        func.extract('month', Order.order_date)).all()

        result = [{
            'period': f"{int(item.year)}/{int(item.month):02d}",
            'total': float(item.total),
            'count': item.count
        } for item in data]

    return jsonify(result)


@financial_reports_bp.route('/api/customer-segments')
@login_required
@permission_required('view_financial_reports')
def api_customer_segments():
    """API برای تقسیم‌بندی مشتریان"""
    # تقسیم‌بندی بر اساس سطح وفاداری
    loyalty_segments = db.session.query(
        Person.person_type,
        func.count(Person.id).label('count'),
        func.sum(Person.total_spent).label('total_spent')).filter(
            Person.person_type == 'customer').group_by(
            Person.person_type).all()

    # تقسیم‌بندی بر اساس تعداد سفارشات
    order_segments = db.session.query(
        func.case([(Person.total_orders == 0, 'بدون سفارش'),
                   (Person.total_orders == 1, 'یک سفارش'),
                   (Person.total_orders.between(2, 5), '2-5 سفارش'),
                   (Person.total_orders.between(6, 10), '6-10 سفارش'),
                   (Person.total_orders
                    > 10, 'بیش از 10 سفارش')]).label('segment'),
        func.count(Person.id).label('count')).filter(
            Person.person_type == 'customer').group_by('segment').all()

    return jsonify({
        'loyalty_segments': [{
            'level': item.customer_level,
            'count': item.count,
            'total_spent': float(item.total_spent)
        } for item in loyalty_segments],
        'order_segments': [{
            'segment': item.segment,
            'count': item.count
        } for item in order_segments]
    })


@financial_reports_bp.route('/export/<report_type>')
@login_required
@permission_required('view_financial_reports')
def export_report(report_type):
    """خروجی اکسل گزارش‌ها"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    wb = Workbook()
    ws = wb.active

    if report_type == 'sales':
        ws.title = "گزارش فروش"
        # اضافه کردن داده‌های فروش
        headers = ['تاریخ', 'تعداد سفارش', 'مبلغ کل', 'میانگین سفارش']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        # داده‌ها
        sales_data = db.session.query(
            func.date(Order.order_date).label('date'),
            func.count(Order.id).label('count'),
            func.sum(Order.final_price).label('total'),
            func.avg(Order.final_price).label('avg')).group_by(
                func.date(Order.order_date)).order_by(
                    func.date(Order.order_date)).all()

        for row, data in enumerate(sales_data, 2):
            ws.cell(row=row, column=1, value=data.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=data.count)
            ws.cell(row=row, column=3, value=float(data.total))
            ws.cell(row=row, column=4, value=float(data.avg))

    elif report_type == 'customers':
        ws.title = "گزارش مشتریان"
        headers = ['نام', 'سطح', 'تعداد سفارش', 'مبلغ کل', 'تاریخ ثبت‌نام']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        customers = Person.query.filter_by(person_type='customer').order_by(desc(Person.total_spent)).all()
        for row, customer in enumerate(customers, 2):
            ws.cell(row=row, column=1, value=customer.full_name)
            ws.cell(row=row, column=2, value=customer.person_type)
            ws.cell(row=row, column=3, value=customer.total_orders)
            ws.cell(row=row, column=4, value=float(customer.total_spent))
            ws.cell(row=row,
                    column=5,
                    value=customer.registration_date.strftime('%Y-%m-%d'))

    # ذخیره فایل
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=
        f'report_{report_type}_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@financial_reports_bp.route('/mechanic-commission/<int:mechanic_id>')
@login_required
@permission_required('view_financial_reports')
def mechanic_commission(mechanic_id):
    """
    گزارش مالی مکانیک خاص
    """
    mechanic = Person.query.filter_by(id=mechanic_id, person_type='mechanic').first_or_404()
    
    # پارامترهای فیلتر
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # محاسبه بازه زمانی
    if start_date and end_date:
        start = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    else:
        end = datetime.datetime.now()
        start = end - timedelta(days=30)
    
    # سفارشات مکانیک در بازه زمانی
    orders = BotOrder.query.filter(
        BotOrder.mechanic_person_id == mechanic_id,
        BotOrder.created_at.between(start, end)
    ).order_by(BotOrder.created_at.desc()).all()
    
    # آمار سفارشات
    total_orders = len(orders)
    total_amount = sum(order.total_amount for order in orders)
    total_commission = sum(order.commission_amount for order in orders if order.commission_amount)
    
    # لیست آیدی سفارشات مکانیک
    order_ids = [order.id for order in orders]
    
    # تاریخچه تراکنش‌های مالی مربوط به این سفارشات
    transactions = []
    if order_ids:
        transactions = FinancialTransaction.query.filter(
            FinancialTransaction.order_id.in_(order_ids),
            FinancialTransaction.created_at.between(start, end)
        ).order_by(FinancialTransaction.created_at.desc()).all()
    
    # آمار ماهانه
    monthly_stats = []
    current_date = start
    while current_date <= end:
        month_start = datetime.datetime(current_date.year, current_date.month, 1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        month_orders = [o for o in orders if month_start <= o.created_at <= month_end]
        month_total = sum(o.total_amount for o in month_orders)
        month_commission = sum(o.commission_amount for o in month_orders if o.commission_amount)
        
        gregorian_date_obj = datetime.date(current_date.year, current_date.month, 1)
        shamsi_label = jdatetime.date.fromgregorian(date=gregorian_date_obj).strftime("%Y/%m")
        
        monthly_stats.append({
            'month': shamsi_label,
            'orders': len(month_orders),
            'total': month_total,
            'commission': month_commission
        })
        
        current_date = (current_date + timedelta(days=32)).replace(day=1)
    
    return render_template('mechanic_commission.html',
                         mechanic=mechanic,
                         orders=orders,
                         transactions=transactions,
                         monthly_stats=monthly_stats,
                         total_orders=total_orders,
                         total_amount=total_amount,
                         total_commission=total_commission,
                         start_date=start_date,
                         end_date=end_date)


@financial_reports_bp.route('/commission-payments')
@login_required
@permission_required('view_financial_reports')
def commission_payments():
    """
    صفحه پرداخت کمیسیون مکانیک‌ها
    """
    from flask import request
    
    # پارامترهای فیلتر
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    
    # دریافت مکانیک‌های تایید شده
    query = Person.query.filter_by(person_type='mechanic', is_approved=True)
    
    if search:
        query = query.filter(
            db.or_(
                Person.full_name.ilike(f'%{search}%'),
                Person.phone_number.ilike(f'%{search}%')
            )
        )
    
    mechanics = query.all()
    
    # محاسبه آمار برای هر مکانیک
    for mechanic in mechanics:
        # محاسبه تعداد سفارشات
        mechanic.total_orders = BotOrder.query.filter_by(mechanic_person_id=mechanic.id).count()
        
        # محاسبه کمیسیون معوق (سفارشاتی که کمیسیون پرداخت نشده)
        pending_commission = db.session.query(func.sum(BotOrder.commission_amount)).filter(
            BotOrder.mechanic_person_id == mechanic.id,
            BotOrder.commission_amount > 0,
            BotOrder.commission_paid == False
        ).scalar() or 0
        
        mechanic.total_commission = pending_commission
        
        # محاسبه تاریخ آخرین پرداخت
        last_payment = FinancialTransaction.query.filter_by(
            mechanic_id=mechanic.id,
            transaction_type='commission_payment'
        ).order_by(FinancialTransaction.created_at.desc()).first()
        
        if last_payment:
            mechanic.last_payment_date = last_payment.created_at
        else:
            mechanic.last_payment_date = None
    
    # اعمال فیلتر بر اساس کمیسیون
    if status == 'pending':
        mechanics = [m for m in mechanics if m.total_commission > 0]
    elif status == 'paid':
        mechanics = [m for m in mechanics if m.total_commission == 0]
    
    # مرتب‌سازی بر اساس کمیسیون معوق
    mechanics.sort(key=lambda x: x.total_commission, reverse=True)
    
    # صفحه‌بندی دستی
    per_page = 20
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_mechanics = mechanics[start_idx:end_idx]
    
    # ایجاد شیء pagination دستی
    class Pagination:
        def __init__(self, items, page, per_page, total):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page
            self.has_prev = page > 1
            self.has_next = page < self.pages
            self.prev_num = page - 1 if page > 1 else None
            self.next_num = page + 1 if page < self.pages else None
            
            def iter_pages(self):
                return range(1, self.pages + 1)
            
            self.iter_pages = iter_pages
    
    mechanics_pagination = Pagination(paginated_mechanics, page, per_page, len(mechanics))
    
    # آمار کلی
    total_mechanics = len(mechanics)
    total_pending_commission = sum(m.total_commission for m in mechanics)
    
    return render_template('commission_payments.html',
                         mechanics=mechanics_pagination,
                         total_mechanics=total_mechanics,
                         total_pending_commission=total_pending_commission,
                         status=status,
                         search=search)


@financial_reports_bp.route('/pay-commission/<int:mechanic_id>', methods=['POST'])
@login_required
@permission_required('view_financial_reports')
def pay_commission(mechanic_id):
    """
    پرداخت کمیسیون مکانیک (روش قدیمی - حفظ شده برای سازگاری)
    """
    mechanic = Person.query.filter_by(id=mechanic_id, person_type='mechanic').first_or_404()
    amount = request.form.get('amount', type=float)
    payment_method = request.form.get('payment_method', '')
    notes = request.form.get('notes', '')
    
    if not amount or amount <= 0:
        return jsonify({'success': False, 'message': 'مبلغ نامعتبر است'})
    
    if amount > mechanic.total_commission:
        return jsonify({'success': False, 'message': 'مبلغ بیشتر از کمیسیون موجود است'})
    
    try:
        # ایجاد تراکنش مالی
        transaction = FinancialTransaction(
            mechanic_id=mechanic.id,
            amount=amount,
            transaction_type='commission_payment',
            payment_method=payment_method,
            notes=notes,
            created_by=current_user.id
        )
        
        db.session.add(transaction)
        
        # کاهش کمیسیون مکانیک
        mechanic.total_commission -= amount
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'کمیسیون {amount} تومان با موفقیت پرداخت شد'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'خطا در پرداخت کمیسیون'})


@financial_reports_bp.route('/pay-commission-detailed/<int:mechanic_id>', methods=['POST'])
@login_required
@permission_required('view_financial_reports')
def pay_commission_detailed(mechanic_id):
    """
    پرداخت کمیسیون مکانیک با جزئیات کامل
    """
    import os
    from werkzeug.utils import secure_filename
    
    mechanic = Person.query.filter_by(id=mechanic_id, person_type='mechanic').first_or_404()
    
    # دریافت داده‌های فرم
    order_ids = request.form.getlist('order_ids')
    amount = request.form.get('amount', type=float)
    payment_method = request.form.get('payment_method', '')
    card_number = request.form.get('card_number', '')
    check_number = request.form.get('check_number', '')
    tracking_number = request.form.get('tracking_number', '')
    payment_date = request.form.get('payment_date', '')
    notes = request.form.get('notes', '')
    
    # اعتبارسنجی
    if not amount or amount <= 0:
        return jsonify({'success': False, 'message': 'مبلغ نامعتبر است'})
    
    if not payment_method:
        return jsonify({'success': False, 'message': 'روش پرداخت انتخاب نشده است'})
    
    if not tracking_number:
        return jsonify({'success': False, 'message': 'شماره پیگیری وارد نشده است'})
    
    if not payment_date:
        return jsonify({'success': False, 'message': 'تاریخ واریز وارد نشده است'})
    
    # اعتبارسنجی روش پرداخت
    if payment_method == 'bank_transfer' and not card_number:
        return jsonify({'success': False, 'message': 'شماره کارت مبدا برای انتقال بانکی الزامی است'})
    
    if payment_method == 'check' and not check_number:
        return jsonify({'success': False, 'message': 'شماره چک برای پرداخت با چک الزامی است'})
    
    try:
        # آپلود فایل فیش واریز (اگر وجود داشته باشد)
        receipt_filename = None
        if 'receipt_image' in request.files:
            file = request.files['receipt_image']
            if file and file.filename:
                filename = secure_filename(file.filename)
                # ایجاد نام فایل منحصر به فرد
                import uuid
                unique_filename = f"receipt_{uuid.uuid4().hex}_{filename}"
                receipt_filename = unique_filename
                
                # ذخیره فایل
                upload_folder = os.path.join(current_app.static_folder, 'receipts')
                if not os.path.exists(upload_folder):
                    os.makedirs(upload_folder)
                
                file_path = os.path.join(upload_folder, unique_filename)
                file.save(file_path)
        
        # تبدیل تاریخ شمسی به میلادی
        try:
            # تبدیل تاریخ شمسی به میلادی
            import jdatetime
            payment_date_parts = payment_date.split('/')
            if len(payment_date_parts) == 3:
                year, month, day = map(int, payment_date_parts)
                jdate = jdatetime.date(year, month, day)
                gregorian_date = jdate.togregorian()
                payment_datetime = datetime.datetime.combine(gregorian_date, datetime.time.min)
            else:
                payment_datetime = datetime.datetime.now()
        except Exception:
            payment_datetime = datetime.datetime.now()
        
        # ایجاد تراکنش مالی
        transaction = FinancialTransaction(
            mechanic_id=mechanic.id,
            amount=amount,
            transaction_type='commission_payment',
            payment_method=payment_method,
            tracking_number=tracking_number,
            payment_date=payment_datetime,
            card_number=card_number if payment_method == 'bank_transfer' else None,
            check_number=check_number if payment_method == 'check' else None,
            receipt_image=receipt_filename,
            notes=notes,
            created_by=current_user.id
        )
        
        db.session.add(transaction)
        
        # بروزرسانی سفارشات انتخاب شده (اگر وجود داشته باشند)
        if order_ids:
            for order_id in order_ids:
                # بررسی اینکه آیا سفارش متعلق به این مکانیک است
                bot_order = BotOrder.query.filter_by(id=order_id, mechanic_person_id=mechanic.id).first()
                if bot_order and bot_order.commission_amount > 0:
                    # صفر کردن کمیسیون سفارش
                    bot_order.commission_amount = 0
                    bot_order.commission_paid = True
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'پرداخت کمیسیون {amount:,} تومان با موفقیت ثبت شد'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطا در پرداخت کمیسیون: {str(e)}")
        return jsonify({'success': False, 'message': 'خطا در ثبت پرداخت کمیسیون'})


@financial_reports_bp.route('/api/transaction-details/<int:transaction_id>')
@login_required
@permission_required('view_financial_reports')
def api_transaction_details(transaction_id):
    """
    API برای دریافت جزئیات تراکنش
    """
    transaction = FinancialTransaction.query.get_or_404(transaction_id)
    
    # تبدیل نوع تراکنش به فارسی
    transaction_type_fa = {
        'commission_payment': 'پرداخت کمیسیون',
        'order_payment': 'پرداخت سفارش',
        'refund': 'بازگشت وجه'
    }.get(transaction.transaction_type, transaction.transaction_type)
    
    # تبدیل روش پرداخت به فارسی
    payment_method_fa = {
        'bank_transfer': 'انتقال بانکی',
        'cash': 'نقدی',
        'check': 'چک',
        'other': 'سایر'
    }.get(transaction.payment_method, transaction.payment_method)
    
    # تبدیل تاریخ‌ها به شمسی
    import jdatetime
    if transaction.created_at:
        jdate = jdatetime.datetime.fromgregorian(datetime=transaction.created_at)
        created_at_fa = jdate.strftime('%Y/%m/%d %H:%M')
    else:
        created_at_fa = '-'
    
    if transaction.payment_date:
        jdate = jdatetime.datetime.fromgregorian(datetime=transaction.payment_date)
        payment_date_fa = jdate.strftime('%Y/%m/%d')
    else:
        payment_date_fa = '-'
    
    return jsonify({
        'success': True,
        'transaction': {
            'id': transaction.id,
            'transaction_type': transaction.transaction_type,
            'transaction_type_fa': transaction_type_fa,
            'amount': transaction.amount,
            'payment_method': transaction.payment_method,
            'payment_method_fa': payment_method_fa,
            'tracking_number': transaction.tracking_number,
            'payment_date': transaction.payment_date.isoformat() if transaction.payment_date else None,
            'payment_date_fa': payment_date_fa,
            'card_number': transaction.card_number,
            'check_number': transaction.check_number,
            'receipt_image': transaction.receipt_image,
            'notes': transaction.notes,
            'created_at': transaction.created_at.isoformat() if transaction.created_at else None,
            'created_at_fa': created_at_fa,
            'mechanic_name': transaction.mechanic.full_name if transaction.mechanic else None
        }
    })


@financial_reports_bp.route('/transaction-history')
@login_required
@permission_required('view_financial_reports')
def transaction_history():
    """
    تاریخچه تراکنش‌های مالی
    """
    # پارامترهای فیلتر
    transaction_type = request.args.get('type', '')
    mechanic_id = request.args.get('mechanic_id', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    page = request.args.get('page', 1, type=int)
    
    query = FinancialTransaction.query
    
    if transaction_type:
        query = query.filter(FinancialTransaction.transaction_type == transaction_type)
    
    if mechanic_id:
        query = query.filter(FinancialTransaction.mechanic_id == mechanic_id)
    
    if start_date:
        try:
            # تبدیل تاریخ شمسی به میلادی
            import jdatetime
            start_date_parts = start_date.split('/')
            if len(start_date_parts) == 3:
                year, month, day = map(int, start_date_parts)
                jdate = jdatetime.date(year, month, day)
                gregorian_date = jdate.togregorian()
                start = datetime.datetime.combine(gregorian_date, datetime.time.min)
                query = query.filter(FinancialTransaction.created_at >= start)
        except (ValueError, TypeError):
            pass
    
    if end_date:
        try:
            # تبدیل تاریخ شمسی به میلادی
            import jdatetime
            end_date_parts = end_date.split('/')
            if len(end_date_parts) == 3:
                year, month, day = map(int, end_date_parts)
                jdate = jdatetime.date(year, month, day)
                gregorian_date = jdate.togregorian()
                end = datetime.datetime.combine(gregorian_date, datetime.time.min) + timedelta(days=1)
                query = query.filter(FinancialTransaction.created_at < end)
        except (ValueError, TypeError):
            pass
    
    transactions = query.order_by(FinancialTransaction.created_at.desc()).paginate(
        page=page, per_page=50, error_out=False
    )
    
    # آمار کلی
    total_transactions = FinancialTransaction.query.count()
    total_amount = db.session.query(func.sum(FinancialTransaction.amount)).scalar() or 0
    
    # دریافت لیست مکانیک‌ها برای فیلتر
    mechanics = Person.query.filter_by(person_type='mechanic', is_approved=True).all()
    
    return render_template('transaction_history.html',
                         transactions=transactions,
                         mechanics=mechanics,
                         total_transactions=total_transactions,
                         total_amount=total_amount,
                         transaction_type=transaction_type,
                         mechanic_id=mechanic_id,
                         start_date=start_date,
                         end_date=end_date)


@financial_reports_bp.route('/customer-orders-history/<int:customer_id>')
@login_required
@permission_required('view_financial_reports')
def customer_orders_history(customer_id):
    """
    گزارش تاریخچه کامل سفارشات یک مشتری خاص
    شامل سفارشات تلگرام، حضوری و ربات
    """
    from app.models import InStoreOrder, BotOrder
    
    customer = Person.query.filter_by(id=customer_id, person_type='customer').first_or_404()
    
    # گرفتن سفارش‌های حضوری
    instore_orders = customer.instore_orders.order_by(InStoreOrder.created_at.desc()).all()
    # گرفتن سفارش‌های ربات (بر اساس شماره تلفن)
    bot_orders = []
    if customer.phone_number:
        bot_orders = BotOrder.query.filter_by(customer_phone=customer.phone_number).order_by(BotOrder.created_at.desc()).all()
    
    # ترکیب و مرتب‌سازی همه سفارشات
    all_orders = []
    for i in instore_orders:
        all_orders.append({
            'id': i.id,
            'date': i.created_at,
            'total_price': i.total_price,
            'status': i.status,
            'type': 'حضوری',
            'order_type': 'instore',  # برای URL
            'order_obj': i,
            'view_url': None,
            # 'invoice_url': url_for('instore_orders.invoice', order_id=i.id)  # حذف شد
        })
    for b in bot_orders:
        all_orders.append({
            'id': b.id,
            'date': b.created_at,
            'total_price': b.total_amount,
            'status': b.status,
            'type': 'ربات',
            'order_type': 'bot',  # برای URL
            'order_obj': b,
            'view_url': url_for('bot_orders.detail', order_id=b.id),
            'invoice_url': None  # ربات فاکتور ندارد
        })
    all_orders.sort(key=lambda x: x['date'], reverse=True)
    
    # محاسبه آمار
    total_orders = len(all_orders)
    total_amount = sum(order['total_price'] for order in all_orders)
    completed_orders = len([o for o in all_orders if o['status'] in ['تکمیل شده', 'تحویل داده شده', 'پرداخت شده']])
    completed_amount = sum(o['total_price'] for o in all_orders if o['status'] in ['تکمیل شده', 'تحویل داده شده', 'پرداخت شده'])
    
    return render_template('customer_orders_history.html',
                           customer=customer,
                           orders=all_orders,
                           total_orders=total_orders,
                           total_amount=total_amount,
                           completed_orders=completed_orders,
                           completed_amount=completed_amount,
                           title=f"تاریخچه سفارشات {customer.full_name}")


@financial_reports_bp.route('/customer-orders-history/<int:customer_id>/print')
@login_required
@permission_required('view_financial_reports')
def customer_orders_history_print(customer_id):
    """
    پرینت تاریخچه کامل سفارشات یک مشتری
    """
    from app.models import InStoreOrder, BotOrder
    
    customer = Person.query.filter_by(id=customer_id, person_type='customer').first_or_404()
    
    # گرفتن سفارش‌های حضوری
    instore_orders = customer.instore_orders.order_by(InStoreOrder.created_at.desc()).all()
    # گرفتن سفارش‌های ربات (بر اساس شماره تلفن)
    bot_orders = []
    if customer.phone_number:
        bot_orders = BotOrder.query.filter_by(customer_phone=customer.phone_number).order_by(BotOrder.created_at.desc()).all()
    
    # ترکیب و مرتب‌سازی همه سفارشات
    all_orders = []
    for i in instore_orders:
        import json
        try:
            products = json.loads(i.products_info)
        except:
            products = []
        all_orders.append({
            'id': i.id,
            'date': i.created_at,
            'total_price': i.total_price,
            'status': i.status,
            'type': 'حضوری',
            'order_obj': i,
            'items': products
        })
    for b in bot_orders:
        import json
        try:
            items = json.loads(b.order_items)
        except:
            items = []
        all_orders.append({
            'id': b.id,
            'date': b.created_at,
            'total_price': b.total_amount,
            'status': b.status,
            'type': 'ربات',
            'order_obj': b,
            'items': items
        })
    all_orders.sort(key=lambda x: x['date'], reverse=True)
    
    # محاسبه آمار
    total_orders = len(all_orders)
    total_amount = sum(order['total_price'] for order in all_orders)
    completed_orders = len([o for o in all_orders if o['status'] in ['تکمیل شده', 'تحویل داده شده', 'پرداخت شده']])
    completed_amount = sum(o['total_price'] for o in all_orders if o['status'] in ['تکمیل شده', 'تحویل داده شده', 'پرداخت شده'])
    
    return render_template('customer_orders_history_print.html',
                           customer=customer,
                           orders=all_orders,
                           total_orders=total_orders,
                           total_amount=total_amount,
                           completed_orders=completed_orders,
                           completed_amount=completed_amount,
                           title=f"تاریخچه سفارشات {customer.full_name}")


@financial_reports_bp.route('/customer-orders-history/<int:customer_id>/single-order/<order_type>/<int:order_id>/print')
@login_required
@permission_required('view_financial_reports')
def customer_single_order_print(customer_id, order_type, order_id):
    """
    پرینت یک سفارش خاص از مشتری
    """
    import datetime
    from app.models import InStoreOrder, BotOrder, BotOrderItem
    
    customer = Person.query.filter_by(id=customer_id, person_type='customer').first_or_404()
    mechanic = None
    if hasattr(customer, 'telegram_id') and customer.telegram_id:
        mechanic = Person.query.filter_by(telegram_id=customer.telegram_id, person_type='mechanic').first()
    
    if order_type == 'instore':
        order = InStoreOrder.query.get_or_404(order_id)
        import json
        try:
            items = json.loads(order.products_info)
        except:
            items = []
        order_data = {
            'id': order.id,
            'date': order.created_at,
            'total_price': order.total_price,
            'status': order.status,
            'type': 'حضوری',
            'items': items
        }
    elif order_type == 'bot':
        order = BotOrder.query.get_or_404(order_id)
        # استفاده از جدول BotOrderItem به جای JSON
        bot_items = BotOrderItem.query.filter_by(order_id=order.id).all()
        items = []
        for item in bot_items:
            items.append({
                'product_name': item.product_name,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': item.total_price
            })
        order_data = {
            'id': order.id,
            'date': order.created_at,
            'total_price': order.total_amount,
            'status': order.status,
            'type': 'ربات',
            'items': items,
            'customer_address': order.customer_address
        }
    else:
        flash('نوع سفارش نامعتبر است.', 'danger')
        return redirect(url_for('financial_reports.customer_orders_history', customer_id=customer_id))
    
    return render_template('customer_single_order_print.html',
                           customer=customer,
                           mechanic=mechanic,
                           order=order_data,
                           now=datetime.datetime.now(),
                           title=f"فاکتور سفارش {order_data['type']} #{order_data['id']}")
