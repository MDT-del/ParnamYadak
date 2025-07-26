# ---------------------------------------------
# فایل: routes.py
# توضیح: مسیرهای مدیریت مکانیک‌ها - شامل تایید، تنظیم کمیسیون و مشاهده آمار
# ---------------------------------------------

from flask import render_template, request, jsonify, redirect, url_for, flash
from flask_login import login_required, current_user
from app.blueprints.mechanics import mechanics_bp
from app.models import Person, MechanicProfile, BotOrder, db, Notification, Role, AuditLog
from app.utils import shamsi_datetime
import json
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv
import requests


@mechanics_bp.route('/', methods=['GET'])
@login_required
def index():
    """
    صفحه اصلی مدیریت مکانیک‌ها یا خروجی جستجوی AJAX
    """
    if request.args.get('ajax') == '1':
        q = request.args.get('q', '').strip()
        query = Person.query.filter_by(person_type='mechanic')
        if q:
            query = query.filter(
                db.or_(
                    Person.full_name.ilike(f'%{q}%'),
                    Person.phone_number.ilike(f'%{q}%')
                )
            )
        mechanics = query.limit(20).all()
        return {'mechanics': [
            {
                'id': m.id,
                'full_name': m.full_name,
                'phone_number': m.phone_number
            } for m in mechanics
        ]}
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    sort = request.args.get('sort', '')
    
    query = Person.query.filter_by(person_type='mechanic')
    
    # فیلتر وضعیت
    if status == 'approved':
        query = query.join(MechanicProfile).filter(MechanicProfile.is_approved == True)
    elif status == 'pending':
        query = query.outerjoin(MechanicProfile).filter(MechanicProfile.is_approved == False)
    
    # فیلتر جستجو
    if search:
        query = query.filter(
            db.or_(
                Person.full_name.ilike(f'%{search}%'),
                Person.phone_number.ilike(f'%{search}%'),
                Person.telegram_id.ilike(f'%{search}%')
            )
        )
    
    # فیلتر تاریخ
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(Person.registration_date >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Person.registration_date < to_date)
        except ValueError:
            pass
    
    # مرتب‌سازی
    if sort == 'commission':
        query = query.join(MechanicProfile).order_by(MechanicProfile.total_commission.desc())
    elif sort == 'orders':
        query = query.order_by(Person.total_orders.desc())
    else:
        query = query.order_by(Person.registration_date.desc())
    
    mechanics = query.paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('mechanics/index.html', 
                         mechanics=mechanics,
                         status=status)


@mechanics_bp.route('/export')
@login_required
def export():
    """
    خروجی اکسل مکانیک‌ها
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    
    # دریافت فیلترها
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    query = Person.query.filter_by(person_type='mechanic')
    
    # اعمال فیلترها
    if status == 'approved':
        query = query.join(MechanicProfile).filter(MechanicProfile.is_approved == True)
    elif status == 'pending':
        query = query.outerjoin(MechanicProfile).filter(MechanicProfile.is_approved == False)
    
    if search:
        query = query.filter(
            db.or_(
                Person.full_name.ilike(f'%{search}%'),
                Person.phone_number.ilike(f'%{search}%'),
                Person.telegram_id.ilike(f'%{search}%')
            )
        )
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(Person.registration_date >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Person.registration_date < to_date)
        except ValueError:
            pass
    
    mechanics = query.order_by(Person.registration_date.desc()).all()
    
    # ایجاد فایل اکسل
    wb = Workbook()
    ws = wb.active
    ws.title = "مکانیک‌ها"
    
    # تنظیم هدر
    headers = [
        'نام و نام خانوادگی', 'شماره تلفن', 'آیدی تلگرام', 'وضعیت', 
        'درصد کمیسیون', 'تعداد سفارشات', 'مجموع کمیسیون', 'تاریخ ثبت'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # اضافه کردن داده‌ها
    for row, mechanic in enumerate(mechanics, 2):
        ws.cell(row=row, column=1, value=mechanic.full_name)
        ws.cell(row=row, column=2, value=mechanic.phone_number)
        ws.cell(row=row, column=3, value=mechanic.telegram_id)
        ws.cell(row=row, column=4, value='تایید شده' if mechanic.mechanic_profile and mechanic.mechanic_profile.is_approved else 'در انتظار تایید')
        ws.cell(row=row, column=5, value=f"{mechanic.mechanic_profile.commission_percentage}%" if mechanic.mechanic_profile else '0%')
        ws.cell(row=row, column=6, value=mechanic.total_orders)
        ws.cell(row=row, column=7, value=mechanic.mechanic_profile.total_commission if mechanic.mechanic_profile else 0)
        ws.cell(row=row, column=8, value=mechanic.registration_date.strftime('%Y-%m-%d %H:%M') if mechanic.registration_date else '')
    
    # تنظیم عرض ستون‌ها
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # ذخیره فایل
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'mechanics_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@mechanics_bp.route('/<int:person_id>')
@login_required
def detail(person_id):
    """
    نمایش جزئیات مکانیک
    """
    mechanic = Person.query.filter_by(id=person_id, person_type='mechanic').first_or_404()
    return render_template('mechanics/detail.html', mechanic=mechanic, current_user=current_user)


@mechanics_bp.route('/<int:person_id>/approve', methods=['POST'])
@login_required
def approve(person_id):
    """
    تایید مکانیک
    """
    mechanic = Person.query.filter_by(id=person_id, person_type='mechanic').first_or_404()
    commission_percentage = request.form.get('commission_percentage', type=float)
    
    if not commission_percentage or commission_percentage < 0:
        return jsonify({'success': False, 'message': 'درصد کمیسیون نامعتبر است'})
    
    try:
        if not mechanic.mechanic_profile:
            mechanic.mechanic_profile = MechanicProfile()
        mechanic.mechanic_profile.is_approved = True
        mechanic.mechanic_profile.commission_percentage = commission_percentage
        mechanic.mechanic_profile.approved_by = current_user.id
        db.session.commit()
        
        # ارسال پیام تلگرام به مکانیک پس از تایید
        if mechanic.telegram_id:
            from dotenv import load_dotenv
            import os
            import requests
            load_dotenv(os.path.join(os.path.dirname(__file__), '../../../bot/bot_config.env'))
            bot_token = os.getenv('TELEGRAM_BOT_TOKEN')
            if bot_token:
                url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                text = f"🎉 تبریک! ثبت‌نام شما به عنوان مکانیک تایید شد.\n\n💰 درصد کمیسیون شما: {commission_percentage}%\n\nحالا می‌توانید از تمام امکانات ربات استفاده کنید."
                payload = {
                    "chat_id": int(mechanic.telegram_id),
                    "text": text
                }
                import logging
                try:
                    resp = requests.post(url, data=payload, timeout=5)
                    if resp.status_code == 200:
                        logging.info(f"✅ پیام تایید به مکانیک {mechanic.telegram_id} ارسال شد")
                    else:
                        logging.error(f'❌ Telegram API error: {resp.status_code} - {resp.text}')
                except Exception as e:
                    logging.error(f"❌ خطا در ارسال پیام تلگرام: {e}")
        
        # ارسال پیام به ربات پس از تایید مکانیک
        if mechanic.telegram_id:
            try:
                import requests
                # اگر پشت nginx هستیم و روت فوروارد می‌شود:
                bot_notify_url = "https://panel.parnamyadak.ir/api/mechanic_status_notify"  # این آدرس باید در nginx به ربات فوروارد شود
                requests.post(bot_notify_url, json={"telegram_id": int(mechanic.telegram_id), "status": "approved"}, timeout=5)
            except Exception as e:
                import logging
                logging.error(f"❌ خطا در ارسال اطلاع‌رسانی تایید مکانیک به ربات: {e}")
        
        return jsonify({
            'success': True,
            'message': f'مکانیک {mechanic.full_name} با موفقیت تایید شد'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': 'خطا در تایید مکانیک'})


@mechanics_bp.route('/<int:person_id>/reject', methods=['POST'])
@login_required
def reject(person_id):
    """
    رد درخواست مکانیک
    """
    mechanic = Person.query.filter_by(id=person_id, person_type='mechanic').first_or_404()
    
    try:
        # حذف شخص و پروفایل مکانیک
        if mechanic.mechanic_profile:
            db.session.delete(mechanic.mechanic_profile)
        db.session.delete(mechanic)
        db.session.commit()
        
        # ارسال پیام به ربات (در آینده)
        # send_rejection_message_to_bot(mechanic.telegram_id)
        
        # ارسال پیام به ربات پس از رد مکانیک
        if mechanic.telegram_id:
            try:
                import requests
                # اگر پشت nginx هستیم و روت فوروارد می‌شود:
                bot_notify_url = "https://panel.parnamyadak.ir/api/mechanic_status_notify"  # این آدرس باید در nginx به ربات فوروارد شود
                requests.post(bot_notify_url, json={"telegram_id": int(mechanic.telegram_id), "status": "rejected"}, timeout=5)
            except Exception as e:
                import logging
                logging.error(f"❌ خطا در ارسال اطلاع‌رسانی رد مکانیک به ربات: {e}")
        
        return jsonify({
            'success': True,
            'message': 'مکانیک رد شد'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'خطا در رد مکانیک'})


@mechanics_bp.route('/<int:person_id>/update_commission', methods=['POST'])
@login_required
def update_commission(person_id):
    """
    بروزرسانی درصد کمیسیون
    """
    mechanic = Person.query.filter_by(id=person_id, person_type='mechanic').first_or_404()
    commission_percentage = request.form.get('commission_percentage', type=float)
    
    if not commission_percentage or commission_percentage < 0:
        return jsonify({'success': False, 'message': 'درصد کمیسیون نامعتبر است'})
    
    try:
        if mechanic.mechanic_profile:
            mechanic.mechanic_profile.commission_percentage = commission_percentage
            db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'درصد کمیسیون بروزرسانی شد'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': 'خطا در بروزرسانی کمیسیون'})


@mechanics_bp.route('/api/mechanics')
@login_required
def api_mechanics():
    """
    API برای دریافت لیست مکانیک‌ها (برای ربات)
    """
    mechanics = Person.query.filter_by(person_type='mechanic').join(MechanicProfile).filter(MechanicProfile.is_approved==True).all()
    
    result = []
    for mechanic in mechanics:
        result.append({
            'id': mechanic.id,
            'telegram_id': mechanic.telegram_id,
            'full_name': mechanic.full_name,
            'phone_number': mechanic.phone_number,
            'commission_percentage': mechanic.mechanic_profile.commission_percentage if mechanic.mechanic_profile else 0
        })
    
    return jsonify(result)


@mechanics_bp.route('/api/register', methods=['POST'])
def api_register_mechanic():
    """
    API برای ثبت‌نام مکانیک از ربات
    حالا عکس جواز کسب را هم می‌پذیرد (اختیاری)
    """
    if request.content_type and request.content_type.startswith('multipart/form-data'):
        data = request.form
        file = request.files.get('business_license_image')
    else:
        data = request.get_json() or {}
        file = None

    telegram_id = data.get('telegram_id')
    full_name = data.get('full_name')
    phone_number = data.get('phone_number')
    card_number = data.get('card_number')
    sheba_number = data.get('sheba_number')
    shop_address = data.get('shop_address')
    business_license = data.get('business_license')
    business_license_image = None
    username = data.get('username')

    # ذخیره فایل اگر وجود داشت
    if file:
        import os
        from werkzeug.utils import secure_filename
        upload_dir = os.path.join('app', 'static', 'mechanic_licenses')
        os.makedirs(upload_dir, exist_ok=True)
        filename = secure_filename(f"{telegram_id}_license_{file.filename}")
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)
        business_license_image = os.path.relpath(file_path, os.path.join('app', 'static')).replace('\\', '/').replace('\\', '/')

    if not all([telegram_id, full_name, phone_number]):
        return jsonify({'success': False, 'message': 'اطلاعات ناقص'})

    try:
        # بررسی تکراری نبودن
        existing_mechanic = Person.query.filter_by(telegram_id=telegram_id, person_type='mechanic').first()
        if existing_mechanic:
            return jsonify({'success': False, 'message': 'شما قبلاً ثبت‌نام کرده‌اید'})
        # بررسی تکراری نبودن شماره تلفن
        existing_phone = Person.query.filter_by(phone_number=phone_number, person_type='mechanic').first()
        if existing_phone:
            # اگر شماره تلفن قبلاً ثبت شده، telegram_id را به آن اضافه کن (در صورت نبود)
            if not existing_phone.telegram_id:
                existing_phone.telegram_id = telegram_id
                db.session.commit()
                return jsonify({'success': True, 'message': 'مکانیک با موفقیت به‌روزرسانی شد', 'mechanic_id': existing_phone.id})
            else:
                return jsonify({'success': False, 'message': 'این شماره تلفن قبلاً ثبت شده است.'})
        # ایجاد شخص جدید
        person = Person(
            telegram_id=telegram_id,
            full_name=full_name,
            phone_number=phone_number,
            person_type='mechanic',
            address=shop_address,
            username=username
        )
        db.session.add(person)
        db.session.flush()

        # ایجاد پروفایل مکانیک
        mechanic_profile = MechanicProfile(
            person_id=person.id,
            card_number=card_number,
            sheba_number=sheba_number,
            shop_address=shop_address,
            business_license=business_license,
            business_license_image=business_license_image
        )
        db.session.add(mechanic_profile)
        db.session.commit()
        
        # ایجاد اعلان برای ادمین‌ها
        try:
            from app.models import Notification, Role
            admin_role = Role.query.filter_by(name='admin').first()
            if admin_role:
                notification = Notification(
                    message=f'مکانیک جدید {full_name} ثبت‌نام کرده است',
                    role_id=admin_role.id
                )
                db.session.add(notification)
                db.session.commit()
                logging.info(f"نوتیفیکیشن برای مکانیک جدید {full_name} ایجاد شد")
        except Exception as e:
            # اگر خطایی در ایجاد اعلان رخ داد، آن را لاگ کنیم اما ثبت‌نام را متوقف نکنیم
            import logging
            logging.error(f"Error creating admin notification for new mechanic: {e}")

        return jsonify({
            'success': True,
            'message': 'درخواست ثبت‌نام با موفقیت ارسال شد',
            'id': person.id
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'خطا در ثبت‌نام'}) 


@mechanics_bp.route('/health', methods=['GET'])
def health_check():
    """بررسی سلامت API"""
    return jsonify({'status': 'healthy', 'service': 'mechanics_api'})

@mechanics_bp.route('/api/status', methods=['GET'])
def mechanic_status():
    """
    دریافت وضعیت مکانیک بر اساس telegram_id
    ورودی: ?telegram_id=...
    خروجی: {"status": "pending"/"approved"/"rejected", ...}
    """
    telegram_id = request.args.get('telegram_id', type=int)
    if not telegram_id:
        return jsonify({'success': False, 'message': 'telegram_id الزامی است'}), 400
    mechanic = Person.query.filter_by(telegram_id=telegram_id, person_type='mechanic').first()
    if not mechanic:
        return jsonify({'success': False, 'message': 'مکانیک یافت نشد'}), 404
    
    status = 'pending'
    commission_percentage = 0
    if mechanic.mechanic_profile:
        if mechanic.mechanic_profile.is_approved:
            status = 'approved'
        elif mechanic.mechanic_profile.is_rejected:
            status = 'rejected'
        commission_percentage = mechanic.mechanic_profile.commission_percentage or 0

    return jsonify({
        'success': True,
        'status': status,
        'commission_percentage': commission_percentage,
        'mechanic_id': mechanic.id,
        'full_name': mechanic.full_name,
        'phone_number': mechanic.phone_number
    })

@mechanics_bp.route('/api/user/status', methods=['GET'])
def user_status():
    """
    بررسی وضعیت کاربر (مکانیک یا مشتری) بر اساس telegram_id
    ورودی: ?telegram_id=...
    خروجی: {"success": true, "status": "approved", "role": "mechanic"/"customer", ...}
    """
    # استفاده از مدل Person که شامل هم مشتری و هم مکانیک است
    
    telegram_id = request.args.get('telegram_id', type=int)
    if not telegram_id:
        return jsonify({'success': False, 'message': 'telegram_id الزامی است'}), 400
    
    # بررسی هر دو نوع شخص (مکانیک و مشتری)
    person = Person.query.filter_by(telegram_id=telegram_id).first()
    if person:
        if person.person_type == 'mechanic':
            status = 'pending'
            commission_percentage = 0
            if person.mechanic_profile:
                if person.mechanic_profile.is_approved:
                    status = 'approved'
                elif person.mechanic_profile.is_rejected:
                    status = 'rejected'
                commission_percentage = person.mechanic_profile.commission_percentage or 0

            return jsonify({
                'success': True,
                'status': status,
                'role': 'mechanic',
                'commission_percentage': commission_percentage,
                'user_id': person.id,
                'full_name': person.full_name,
                'phone_number': person.phone_number
            })
        elif person.person_type == 'customer':
            return jsonify({
                'success': True,
                'status': 'approved',  # مشتریان به طور خودکار تایید شده هستند
                'role': 'customer',
                'user_id': person.id,
                'full_name': person.full_name,
                'phone_number': person.phone_number
            })
    
    # اگر هیچ‌کدام نبود، کاربر جدید است - به عنوان مشتری ثبت کن
    try:
        new_customer = Person(
            full_name=f'کاربر {telegram_id}',
            phone_number=None,  # بعداً توسط کاربر تکمیل می‌شود
            telegram_id=telegram_id,
            person_type='customer'
        )
        db.session.add(new_customer)
        db.session.commit()

        import logging
        logging.info(f"[USER_STATUS] Auto-registered new customer with telegram_id: {telegram_id}")

        return jsonify({
            'success': True,
            'status': 'approved',  # مشتریان جدید خودکار تایید می‌شوند
            'role': 'customer',
            'user_id': new_customer.id,
            'full_name': new_customer.full_name,
            'phone_number': new_customer.phone_number
        })

    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"[USER_STATUS] Error auto-registering customer {telegram_id}: {e}")
        return jsonify({'success': False, 'message': 'کاربر یافت نشد'}), 404


@mechanics_bp.route('/<int:person_id>/edit', methods=['GET', 'POST'])
@login_required
# فقط ادمین‌ها
def edit_mechanic(person_id):
    from app import db
    mechanic = Person.query.filter_by(id=person_id, person_type='mechanic').first_or_404()
    if not current_user.has_role('admin'):
        flash('دسترسی فقط برای ادمین!', 'danger')
        return redirect(url_for('mechanics.detail', person_id=person_id))
    if request.method == 'POST':
        mechanic.full_name = request.form.get('full_name')
        mechanic.phone_number = request.form.get('phone_number')
        mechanic.address = request.form.get('shop_address')
        
        if mechanic.mechanic_profile:
            mechanic.mechanic_profile.card_number = request.form.get('card_number')
            mechanic.mechanic_profile.sheba_number = request.form.get('sheba_number')
            mechanic.mechanic_profile.shop_address = request.form.get('shop_address')
            mechanic.mechanic_profile.business_license = request.form.get('business_license')
        
        db.session.commit()
        # ثبت لاگ ویرایش
        log = AuditLog(user_id=current_user.id, action='ویرایش اطلاعات مکانیک', description=f'ویرایش اطلاعات مکانیک {mechanic.full_name} (ID: {mechanic.id}) توسط {current_user.name}', ip_address=request.remote_addr)
        db.session.add(log)
        db.session.commit()
        flash('اطلاعات مکانیک با موفقیت ویرایش شد.', 'success')
        return redirect(url_for('mechanics.detail', person_id=person_id))
    return render_template('mechanics/edit.html', mechanic=mechanic)